import openpyxl
from openpyxl.styles import Alignment
from pathlib import Path

# Mapeo de celdas en la plantilla
MAPA_CELDAS = {
    "contratista": "B7",
    "contrato": "I7", 
    "descripcion": "B8",
    "monto": "C13",
    "plazo_dias": "F13",
    "anexos_inicio": "B29",
}

def load_excel(file_path):
    """Cargar archivo Excel de forma segura"""
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
    
    return openpyxl.load_workbook(file_path)

def write_data(sheet, data_dict):
    """
    Escribir datos en la plantilla Excel
    Maneja celdas fusionadas correctamente
    """
    
    def _write_cell(cell_coord, value):
        """Escribir en celda manejando fusiones"""
        # Verificar si la celda está fusionada
        for merged_range in sheet.merged_cells.ranges:
            if cell_coord in merged_range:
                # Desfusionar temporalmente
                sheet.unmerge_cells(str(merged_range))
                sheet[cell_coord] = value
                # Volver a fusionar
                sheet.merge_cells(str(merged_range))
                return
        
        # Celda normal
        sheet[cell_coord] = value

    # CONTRATISTA
    if data_dict.get("contratista"):
        _write_cell(MAPA_CELDAS["contratista"], f"Contratista: {data_dict['contratista']}")

    # NÚMERO DE CONTRATO
    if data_dict.get("contrato"):
        numero = str(data_dict["contrato"])
        if not numero.startswith("64"):
            numero = f"64{numero}"
        _write_cell(MAPA_CELDAS["contrato"], f"NO. {numero}")

    # DESCRIPCIÓN
    if data_dict.get("descripcion"):
        _write_cell(MAPA_CELDAS["descripcion"], f"Descripción del contrato: {data_dict['descripcion']}")

    # MONTO
    if data_dict.get("monto"):
        monto = str(data_dict["monto"]).strip()
        if not monto.startswith("$"):
            monto = f"${monto}"
        _write_cell(MAPA_CELDAS["monto"], monto)

    # PLAZO
    if data_dict.get("plazo_dias"):
        _write_cell(MAPA_CELDAS["plazo_dias"], str(data_dict["plazo_dias"]))

    # ANEXOS
    if data_dict.get("anexos") and isinstance(data_dict["anexos"], list):
        col_inicio = openpyxl.utils.column_index_from_string(MAPA_CELDAS["anexos_inicio"][0])
        fila_inicio = int(MAPA_CELDAS["anexos_inicio"][1:])
        
        for i, anexo in enumerate(data_dict["anexos"]):
            if i >= 50:  # Límite de anexos
                break
                
            celda = sheet.cell(row=fila_inicio + i, column=col_inicio)
            celda.value = anexo
            celda.alignment = Alignment(horizontal="left", vertical="center")

def save_excel(workbook, output_path):
    """Guardar archivo Excel de forma segura"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    workbook.save(output_path)
    print(f"✅ Excel guardado: {output_path}")